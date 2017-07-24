# -*- coding:utf-8 -*-

import os
import copy
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook


class Reader(object):
    def __init__(self, filename):
        self.workbook = load_workbook(filename)
        self.sheet_count = len(self.workbook.worksheets)
        self.buy_sheet = None
        self.pay_sheet = None
        self.buy_title_row = None
        self.pay_title_row = None

    def read(self):
        for i in range(0, self.sheet_count, 2):
            self.read_sheet(i)

    def read_sheet(self, index):
        self.buy_sheet = self.workbook.worksheets[index]
        self.pay_sheet = self.workbook.worksheets[index + 1]
        buy_sheet = self.buy_sheet
        pay_sheet = self.pay_sheet

        buyno = buy_sheet.columns.next()
        rows = []
        for row in buy_sheet.rows:
            rows.append(row)
        row_num = 0
        self.buy_title_row = self.tran(rows[0])
        buylist = []
        for no in buyno[1:]:
            row_num += 1
            row = rows[row_num]
            no_value = no.value
            if no_value is None:
                no_value = ''
            temp = {'buyno': no_value.split(',')[0].strip(),
                    'row': self.tran(row)}
            temp['name'] = temp['row'][3].strip()
            temp['time'] = temp['row'][5]
            temp['value'] = temp['row'][7]
            buylist.append(temp)

        payno = self.pay_sheet.columns.next()
        rows = []
        for row in pay_sheet.rows:
            rows.append(row)
        row_num = 0
        self.pay_title_row = self.tran(rows[0])
        paylist = []
        for no in payno[1:]:
            row_num += 1
            row = rows[row_num]
            no_value = no.value
            if no_value is None:
                no_value = ''
            temp = {'payno': no_value.split(',')[0].strip(),
                    'row': self.tran(row)}
            temp['name'] = temp['row'][5].strip()
            temp['time'] = temp['row'][3]
            temp['value'] = temp['row'][9] - temp['row'][13]
            paylist.append(temp)

        buylist = self.distinct(buylist)
        paylist = self.distinct(paylist)

        buy_emptyrow = copy.copy(self.buy_title_row)
        for i in range(0, len(buy_emptyrow)):
            buy_emptyrow[i] = ''

        pay_emptyrow = copy.copy(self.pay_title_row)
        for i in range(0, len(pay_emptyrow)):
            pay_emptyrow[i] = ''

        inner_list = []
        left_list = []
        right_list = []

        for buy in buylist:
            temp = {'buy': buy, 'pay': {'payno': '', 'name': '', 'time': '', 'value': '', 'row': pay_emptyrow}}
            for pay in paylist:
                if buy['buyno'].__class__ == pay['payno'].__class__ and \
                                buy['buyno'] == pay['payno']:
                    temp['pay'] = pay
                    inner_list.append(temp)
                    break
            left_list.append(temp)

        for pay in paylist:
            temp = {'buy': {'buy': '', 'name': '', 'time': '', 'value': '', 'row': buy_emptyrow}, 'pay': pay}
            for buy in buylist:
                if buy['buyno'].__class__ == pay['payno'].__class__ and \
                                buy['buyno'] == pay['payno']:
                    temp['pay'] = pay
                    break
            right_list.append(temp)

        self.save('out' + str(index) + '.xlsx', '订单与支付都有'.decode('utf8'), inner_list)
        self.save('out' + str(index) + '.xlsx', '所有订单+支付'.decode('utf8'), left_list)
        self.save('out' + str(index) + '.xlsx', '订单+所有支付'.decode('utf8'), right_list)

    def distinct(self, datalist):
        newlist = []
        cur_row_num = 0
        total_len = datalist.__len__()
        while True:
            row = datalist[cur_row_num]
            newlist.append(row)
            for row_num in range(total_len - 1, cur_row_num, -1):
                rowdata = datalist[row_num]
                if row['name'].__class__ == rowdata['name'].__class__ \
                        and row['name'].strip() == rowdata['name'].strip() \
                        and row['time'].__class__ == rowdata['time'].__class__:
                    try:
                        if row['time'].__class__ == unicode \
                                and abs((datetime.datetime.strptime(row['time'], '%Y-%m-%d %H:%M:%S') -
                                             datetime.datetime.strptime(rowdata['time'],
                                                                        '%Y-%m-%d %H:%M:%S')).days) <= 1:
                            try:
                                row['value'] += rowdata['value']
                            except Exception, e:
                                pass
                            datalist.remove(rowdata)
                        if row['time'].__class__ == datetime \
                                and abs((row['time'] - rowdata['time']).days) <= 1:
                            try:
                                row['value'] += rowdata['value']
                            except Exception, e:
                                pass
                            datalist.remove(rowdata)
                    except Exception, e:
                        pass

            cur_row_num += 1
            total_len = datalist.__len__()
            if cur_row_num >= total_len:
                break

        return newlist

    def save(self, filename, sheetname, outlist):
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
            wb.remove(wb.worksheets[0])
        sheet = wb.create_sheet(sheetname)

        col_i = 1
        row_i = 1
        for title in self.buy_title_row:
            sheet.cell(column=col_i, row=row_i).value = title
            col_i += 1
        sheet.cell(column=col_i, row=row_i).value = '订单总额'
        col_i += 1

        for title in self.pay_title_row:
            sheet.cell(column=col_i, row=row_i).value = title
            col_i += 1
        sheet.cell(column=col_i, row=row_i).value = '支付总额'

        col_i = 1
        row_i += 1

        for item in outlist:
            for ce in item['buy']['row']:
                sheet.cell(column=col_i, row=row_i).value = ce
                col_i += 1
            sheet.cell(column=col_i, row=row_i).value = item['buy']['value']
            col_i += 1

            for ce in item['pay']['row']:
                sheet.cell(column=col_i, row=row_i).value = ce
                col_i += 1
            sheet.cell(column=col_i, row=row_i).value = item['pay']['value']

            col_i = 1
            row_i += 1

        wb.save(filename)
        wb.close()

    def tran(self, row):
        ret = []
        for ce in row:
            ret.append(ce.value)
        return ret


if __name__ == "__main__":
    reader = Reader('in.xlsx')
    reader.read()
